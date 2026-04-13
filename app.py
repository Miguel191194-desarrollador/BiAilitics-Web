# app.py
from flask import Flask, render_template, request, redirect, url_for, flash
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
import io
import threading
import base64
import tempfile
import os
import logging
import requests

# Load environment variables
try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'dimensa-secret-key-2024')

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

# ===== Config =====
GAS_WEBHOOK_URL = os.getenv("GAS_WEBHOOK_URL")
MAIL_TO_ADMIN   = os.getenv("MAIL_TO_ADMIN")
FORCE_SYNC_SEND = os.getenv("FORCE_SYNC_SEND", "false").lower() in ("1", "true", "yes")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_CLIENTE_TEMPLATE = os.path.join(BASE_DIR, "Copia de Alta de Cliente.xlsx")
EXCEL_PLANTAS_TEMPLATE = os.path.join(BASE_DIR, "Copia de Alta de Plantas.xlsx")

# ===== Routes =====
@app.route('/', methods=['GET'])
def formulario():
    return render_template('formulario.html')

@app.route('/plantas', methods=['POST', 'GET'])
def plantas():
    if request.method == 'GET':
        return redirect(url_for('formulario'))
    datos_cliente = request.form.to_dict()
    return render_template('plantas.html', datos_cliente=datos_cliente)

@app.route('/guardar', methods=['POST'])
def guardar():
    form_data = request.form.to_dict()
    data = form_data # In reality, all data comes in form_data because of hidden inputs

    # Signature processing
    firma_base64 = data.get('firma_cliente')
    firma_bytes = None
    if firma_base64 and "," in firma_base64:
        try:
            firma_bytes = base64.b64decode(firma_base64.split(",")[1])
        except Exception:
            logging.error("Error decoding signature")

    # Minimum validation: at least one plant name
    if not data.get('planta_nombre_1'):
        flash('⚠️ Debes rellenar al menos los datos de la primera planta.')
        return render_template('plantas.html', datos_cliente=form_data)

    # Generate Excels
    try:
        excel_cliente = crear_excel_en_memoria(data, firma_bytes)
        excel_plantas = crear_excel_plantas_en_memoria(data)
    except Exception as e:
        logging.exception("❌ Error generating Excels")
        flash(f'Error generando documentación: {e}')
        return render_template('plantas.html', datos_cliente=form_data)

    if not GAS_WEBHOOK_URL:
        logging.error("❌ GAS_WEBHOOK_URL is not configured")
        flash('Error de configuración en el servidor.')
        return render_template('gracias.html')

    nombre_client_clean = data.get('nombre', 'cliente').replace("/", "-")
    correo_comercial = data.get('correo_comercial')

    if FORCE_SYNC_SEND:
        ok, detalle = enviar_un_correo_con_dos_adjuntos(excel_cliente, excel_plantas, correo_comercial, nombre_client_clean)
        if not ok:
            logging.error(f"Sync send failed: {detalle}")
    else:
        threading.Thread(
            target=_thread_enviar_unico,
            args=(excel_cliente, excel_plantas, correo_comercial, nombre_client_clean),
            daemon=True
        ).start()
    
    return render_template("gracias.html")

def _thread_enviar_unico(archivo1, archivo2, correo, nombre):
    try:
        ok, detalle = enviar_un_correo_con_dos_adjuntos(archivo1, archivo2, correo, nombre)
        if ok:
            logging.info("✅ Successfully sent documentation for %s", nombre)
        else:
            logging.error("❌ Failed to send documentation: %s", detalle)
    except Exception as e:
        logging.exception("❌ Exception in sending thread: %s", e)

# ===== Excel Functions =====
def crear_excel_en_memoria(data, firma_bytes=None):
    wb = load_workbook(EXCEL_CLIENTE_TEMPLATE)
    ws = wb["FICHA CLIENTE"]

    # Mapping fields from form to Excel cells
    mapping = {
        "B4": "nombre", "B5": "nif", "D5": "telefono_general",
        "B6": "email_general", "D6": "web", "B7": "direccion",
        "D7": "cp", "B8": "poblacion", "D8": "provincia",
        "B13": "forma_pago", "B18": "compras_nombre", "D18": "compras_telefono",
        "B19": "compras_email", "B22": "contabilidad_nombre", "D22": "contabilidad_telefono",
        "B24": "contabilidad_email", "B27": "facturacion_nombre", "D27": "facturacion_telefono",
        "B29": "facturacion_email", "B32": "descarga_nombre", "D32": "descarga_telefono",
        "B34": "descarga_email", "C38": "contacto_documentacion", "C39": "contacto_devoluciones",
        "B43": "sepa_nombre_banco", "B44": "sepa_domicilio_banco", "B45": "sepa_cp",
        "B46": "sepa_poblacion", "B47": "sepa_provincia", "B48": "iban_completo"
    }

    for cell, field in mapping.items():
        ws[cell] = data.get(field, "")

    if firma_bytes:
        import uuid
        tmp_filename = f"firma_temp_{uuid.uuid4().hex}.png"
        tmp_path = os.path.join(BASE_DIR, tmp_filename)
        with open(tmp_path, "wb") as tmp:
            tmp.write(firma_bytes)
        try:
            img = ExcelImage(tmp_path)
            img.width = 180
            img.height = 70
            ws.add_image(img, "B49")
        finally:
            try:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
            except Exception:
                pass

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

def crear_excel_plantas_en_memoria(data):
    wb = load_workbook(EXCEL_PLANTAS_TEMPLATE)
    ws = wb["Plantas"]

    columnas = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]
    fields = [
        "planta_nombre_{}", "planta_direccion_{}", "planta_cp_{}", "planta_poblacion_{}",
        "planta_provincia_{}", "planta_telefono_{}", "planta_email_{}", "planta_horario_{}",
        "planta_observaciones_{}", "planta_contacto_nombre_{}", "planta_contacto_telefono_{}",
        "planta_contacto_email_{}"
    ]

    for i in range(1, 11):
        fila = 3 + i
        if not data.get(f"planta_nombre_{i}"):
            continue
        
        for idx, field_pattern in enumerate(fields):
            ws[f"{columnas[idx]}{fila}"] = data.get(field_pattern.format(i), "")

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

# ===== Webhook Logic =====
def enviar_un_correo_con_dos_adjuntos(archivo_cliente, archivo_plantas, correo_comercial, nombre_cliente):
    if not GAS_WEBHOOK_URL:
        return False, "Falta GAS_WEBHOOK_URL"

    to_csv = _build_recipients(correo_comercial)
    subject = f"Alta de cliente: {nombre_cliente}"
    body_html = construir_body_html(nombre_cliente)

    att1 = _encode_attachment(archivo_cliente, f"Alta_Cliente_{nombre_cliente}.xlsx")
    att2 = _encode_attachment(archivo_plantas, f"Plantas_{nombre_cliente}.xlsx")

    payload = {
        "to": to_csv,
        "subject": subject,
        "html": body_html,
        "attachments": [att1, att2]
    }

    try:
        r = requests.post(GAS_WEBHOOK_URL, json=payload, timeout=30)
        return r.status_code == 200, r.text
    except Exception as e:
        return False, str(e)

def _build_recipients(correo_comercial):
    dest = ['tesoreria@dimensasl.com']
    if correo_comercial and "@" in correo_comercial:
        dest.append(correo_comercial)
    if MAIL_TO_ADMIN and "@" in MAIL_TO_ADMIN:
        dest.append(MAIL_TO_ADMIN)
    return ",".join(list(set(dest)))

def _encode_attachment(bio, filename):
    bio.seek(0)
    b64 = base64.b64encode(bio.getvalue()).decode("utf-8")
    return {
        "filename": filename,
        "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "content": b64
    }

def construir_body_html(nombre_cliente):
    return f"""
    <div style="font-family: Arial, sans-serif; color: #333;">
        <h2>Notificación de Alta de Cliente</h2>
        <p>Se ha registrado un nuevo alta para el cliente: <strong>{nombre_cliente}</strong>.</p>
        <p>Se adjuntan los documentos correspondientes con la ficha de cliente y el detalle de sus plantas.</p>
        <hr>
        <p style="color: #cc0000; font-weight: bold;">⚠️ IMPORTANTE: REENVIAR ESTE CORREO A GERENCIA INDICANDO RIESGO, SECTOR Y SUBSECTOR.</p>
        <p>Un saludo,<br><strong>Departamento de Tesorería</strong></p>
    </div>
    """

if __name__ == '__main__':
    app.run(debug=True)





